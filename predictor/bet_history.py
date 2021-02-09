import os
import numpy as np
import openpyxl
import xlwings as xw

from oddsportal_scraper import *


def bet_history(prob, odds, bookmakers, model_name):
    # Creates file "model.xlsx" if it doesn't exist
    if not os.path.isfile('.\\' + model_name + '.xlsx'):
        wb_history = openpyxl.Workbook()
        sheet_history = wb_history.active
        header = ['date', 'team1', 'team2', 'league_id', 'prob1', 'prob2', 'probTie', 'odd1', 'odd2', 'oddTie',
                  'bookmaker1', 'bookmaker2', 'bookmakerTie', 'outcome1', 'outcome2', 'outcomeTie']
        sheet_history.append(header)
        wb_history.save(model_name + '.xlsx')

    wb_today = openpyxl.load_workbook(r'.\merged_data.xlsx')
    sheet_today = wb_today.active
    wb_history = openpyxl.load_workbook('.\\' + model_name + '.xlsx')
    sheet_history = wb_history.active

    id_history = []
    id_history_without_result = []
    league_without_result = []
    for row in sheet_history.iter_rows(min_row=2):
        id_history.append([row[0].value, row[1].value, row[2].value])
        if row[15].value is None:
            id_history_without_result.append([row[0].value, row[1].value, row[2].value])
            if str(row[3].value) not in league_without_result:
                league_without_result.append(str(row[3].value))
    # ADDS RESULTS FROM YESTERDAY
    # scrapes result data and saves into "data_oddsportal.json"
    oddsportal_scraper(league_without_result, scrape_results=True)
    with open(r'.\data_oddsportal.json') as jsonFile:
        data_oddsportal = json.load(jsonFile)
        for match in data_oddsportal:
            match_id = [data_oddsportal[match]['date'], data_oddsportal[match]['team1'],
                        data_oddsportal[match]['team2']]
            if match_id in id_history_without_result:
                sheet_history.cell(id_history.index(match_id) + 2, 14).value = data_oddsportal[match]['win1']
                sheet_history.cell(id_history.index(match_id) + 2, 15).value = data_oddsportal[match]['win2']
                sheet_history.cell(id_history.index(match_id) + 2, 16).value = data_oddsportal[match]['tie']
    # ADDS OR UPDATES TODAY'S MATCHES
    i = 0
    for row in sheet_today.iter_rows(min_row=1):
        if i == 0:  # gets league id order from "merged_data.xlsx"
            league_id_order = []
            for col in sheet_today.iter_cols(min_col=30):
                league_id_order.append(col[i].value)
        else:
            # identifies league id
            j = 0
            league_id = 0
            for col in sheet_today.iter_cols(min_col=30):
                if col[i].value == '1':
                    league_id = league_id_order[j]
                    break
                j += 1
            id_today = [row[18].value + '-' + row[19].value + '-' + row[20].value, row[0].value, row[1].value]
            if id_today not in id_history:  # appends data if item does not exist
                sheet_history.append(id_today + [league_id] + prob[i - 1].tolist() + odds[i - 1].tolist() + bookmakers[i - 1])
            else:  # updates data if item already exists
                index = id_history.index(id_today)
                sheet_history.cell(index + 2, 5).value = prob[i-1][0]
                sheet_history.cell(index + 2, 6).value = prob[i-1][1]
                sheet_history.cell(index + 2, 7).value = prob[i-1][2]
                sheet_history.cell(index + 2, 8).value = odds[i-1][0]
                sheet_history.cell(index + 2, 9).value = odds[i-1][1]
                sheet_history.cell(index + 2, 10).value = odds[i-1][2]
                sheet_history.cell(index + 2, 11).value = bookmakers[i-1][0]
                sheet_history.cell(index + 2, 12).value = bookmakers[i-1][1]
                sheet_history.cell(index + 2, 13).value = bookmakers[i-1][2]
        i += 1
    wb_history.save(model_name + '.xlsx')
    # EXPANDS FORMULAS
    app = xw.App(visible=False)
    wb = xw.Book(model_name + '.xlsx')
    i = 1
    sheet = wb.sheets.active
    while True:
        cell = 'A' + str(i)
        if sheet[cell].value == datetime.now().strftime('%d-%m-%Y'):
            break
        i += 1
    formula = sheet.range('Q2').formula
    cells_copy = 'Q2:' + 'S' + str(i-1)
    sheet.range(cells_copy).formula = formula

    formula = sheet.range('T2').formula
    cells_copy = 'T2:' + 'V' + str(i-1)
    sheet.range(cells_copy).formula = formula

    formula = sheet.range('AL2').formula
    cells_copy = 'AL2:' + 'AN' + str(i-1)
    sheet.range(cells_copy).formula = formula

    formula = sheet.range('AO2').formula
    cells_copy = 'AO2:' + 'AQ' + str(i-1)
    sheet.range(cells_copy).formula = formula

    sheet.range('AB1:AC1').value = sheet.range('AB3:AC3').value
    sheet.range('AD3:AF3').value = sheet.range('AD1:AF1').value
    simple_evolutionary_optimization(sheet)

    while True:
        cell = 'A' + str(i)
        if sheet[cell].value is None:
            break
        i += 1
    formula = sheet.range('Q2').formula
    cells_copy = 'Q2:' + 'S' + str(i-1)
    sheet.range(cells_copy).formula = formula
    wb.save()
    wb.close()
    app.kill()


def simple_evolutionary_optimization(sheet):
    i = 0
    best_cost = sheet.range('AI2').value
    param1 = sheet.range('AB1').value
    param2 = sheet.range('AC1').value
    threshold = 2**20
    while i <= threshold:
        param1_list = np.random.uniform(low=max(.5, param1/1.05), high=min(2, param1*1.05), size=(45, 1))
        param1_list = np.concatenate((param1_list, np.random.uniform(low=.5, high=2, size=(5, 1)))).tolist()
        param2_list = np.random.uniform(low=max(.5, param2/1.05), high=min(2, param2*1.05), size=(45, 1))
        param2_list = np.concatenate((param2_list, np.random.uniform(low=.5, high=2, size=(5, 1)))).tolist()
        best_cost2 = -9999999999999999999
        for k in range(len(param1_list)):
            sheet.range('AB1').value = param1_list[k][0]
            sheet.range('AC1').value = param2_list[k][0]
            cost = sheet.range('AI2').value
            if cost > best_cost2:
                best_cost2 = cost
                best_param1 = param1_list[k][0]
                best_param2 = param2_list[k][0]
        if best_cost2 > best_cost:
            best_cost = best_cost2
            param1 = best_param1
            param2 = best_param2
            i = 0
        else:
            i += i+1
        print(str(round(i/threshold*100, 0)) + '%')
    sheet.range('AB1').value = param1
    sheet.range('AC1').value = param2
