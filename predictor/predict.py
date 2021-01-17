import copy

import numpy as np
import openpyxl

from neural_net import *


# Loads the model parameters from .mat file (neural network trained in Matlab) and predicts probability
def predict(model):
    # LOADS MODEL DATA
    nn_params, _, _, input_layer_size, hidden_layers_size, num_labels, mu_nn, sigma_nn, mu, sigma = load_model(model)
    # LOADS MATCH DATA (raw)
    match_data = load_match_data()
    if match_data.size == 0:
        return np.array([]), np.array([])
    odds = copy.deepcopy(match_data[:, 5:8])
    match_data[:, 5:8] = 1 / match_data[:, 5:8]
    # NORMALIZES DATA
    match_data = match_data - mu
    match_data = match_data / sigma
    # RESHAPES LAYER VECTOR (this is vector with the layer size of the model)
    layers = np.concatenate((input_layer_size, hidden_layers_size, num_labels), axis=None)
    # RESHAPES PARAMETERS THETA, GAMMA, BETA
    theta, gamma, beta = parameter_reshape(layers, nn_params)
    # CALCULATES PROBABILITY
    prob = forward_propagation(layers, match_data, theta, mu_nn, sigma_nn, gamma, beta)
    return prob, odds


def load_match_data():  # loads data in the correct order of neural network input
    workbook = openpyxl.load_workbook(r'.\merged_data.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    header_order = ['spi1', 'spi2', 'prob1', 'prob2', 'probtie', 'oddsW1', 'oddsW2', 'oddsTie', 'proj_score1',
                    'proj_score2', 'importance1', 'importance2', 'capacity', 'stadium_lat', 'stadium_lon', 'distance',
                    'match_day', 'match_month', 'match_year', 'date', 'dayOfYear', 'dayOfWeek', 'numberBookmarkers',
                    'matchTime', '1818', '1820', '1827', '1832', '1837', '1843', '1844', '1845', '1846', '1849', '1854',
                    '1856', '1859', '1864', '1866', '1869', '1871', '1874', '1879', '1882', '1884', '1947', '1948',
                    '1951', '1952', '1975', '1979', '1983', '2105', '2160', '2411', '2412', '2413', '2414', '2417',
                    '4582', '5641', '9541']
    y = 1
    while sheet.cell(row=y, column=1).value is not None:
        y += 1
    number_matches = y
    if number_matches > 1:
        match_data = np.zeros((number_matches - 2, 62))
    else:
        return np.array([])
    y = 2
    while sheet.cell(row=y, column=1).value is not None:
        x = 1
        while sheet.cell(row=1, column=x).value is not None:
            header = str(sheet.cell(row=1, column=x).value)
            if header in header_order:
                match_data[y - 2][header_order.index(header)] = float(sheet.cell(row=y, column=x).value)
            x += 1
        y += 1
    workbook.close()
    return match_data
