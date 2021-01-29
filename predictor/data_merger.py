import random
from datetime import date

from dictionaries import *
from fivethirtyeight_scraper import *
from oddsportal_scraper import *
from timezonefinder import TimezoneFinder
from math import sin, cos, sqrt, atan2, radians

import json
from openpyxl import Workbook

tf = TimezoneFinder()


# calculates distance from 2 stadium coordinates
def distance_calculator(lat1, lon1, lat2, lon2):
    r = 6373.0  # approximate radius of earth in km
    lat1 = radians(lat1)
    lon1 = radians(lon1)
    lat2 = radians(lat2)
    lon2 = radians(lon2)
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    distance = r * c
    return distance


# Writes "merged_data.json" created by data_merger data into "merged_data.xlsx".
def json_to_excel():
    # READS DATA FROM JSON FILES
    with open(r'.\merged_data.json') as jsonFile:
        match_data = json.load(jsonFile)

    wb = Workbook()
    ws = wb.active
    y = 1
    for match in list(match_data):
        x = 1
        if len(list(match_data[match].keys())) != 67:
            continue
        if y == 1:
            for key in list(match_data[match].keys()):
                ws.cell(row=1, column=x, value=key)
                x += 1
            y += 1
            x = 1
        for item in list(match_data[match].keys()):
            ws.cell(row=y, column=x, value=match_data[match][item])
            x += 1
        y += 1
    wb.save('merged_data.xlsx')


# Gets data from 538 and oddsportal, merges and saves it into "merged_data.json
# Todo: this function is a mess --> clean up, split into more functions (?)
def data_merger():
    # IMPORTS DICTIONARIES
    stadium_capacity, stadium_lat, stadium_lon = stadium_dictionary()
    timezone_dict = timezone_dictionary()
    team_dict = team_dictionary()
    # GETS DATA FROM 538
    save_file_538()
    data_538 = get_today_matches()
    # GETS DATA FROM ODDSPORTAL
    leagueid_list = []
    for match in data_538:
        if data_538[match]['league_id'] not in leagueid_list:
            leagueid_list.append(data_538[match]['league_id'])
    oddsportal_scraper(leagueid_list)
    with open(r'.\data_oddsportal.json') as jsonFile:
        data_oddsportal = json.load(jsonFile)

    # MERGES DATA
    date_regex = re.compile('([0-9]+)-([0-9]+)-([0-9]+)')

    data_538_items = list(data_538)
    random.shuffle(data_538_items)

    for match in data_538_items:
        data_538[match]['capacity'] = stadium_capacity[data_538[match]['team1']]
        data_538[match]['stadium_lat'] = stadium_lat[data_538[match]['team1']]
        data_538[match]['stadium_lon'] = stadium_lon[data_538[match]['team1']]
        data_538[match]['distance'] = str(distance_calculator(float(stadium_lat[data_538[match]['team1']]),
                                                              float(stadium_lon[data_538[match]['team1']]),
                                                              float(stadium_lat[data_538[match]['team2']]),
                                                              float(stadium_lon[data_538[match]['team2']])))

        Pass1 = False
        Pass2 = False
        if data_538[match]['team1'] == 'North Carolina FC':
            if data_538[match]['league_id'] == '2160':
                data_538[match]['team1'] = 'North Carolina'
                Pass1 = True
        if data_538[match]['team2'] == 'North Carolina FC':
            if data_538[match]['league_id'] == '2160':
                data_538[match]['team2'] = 'North Carolina'
                Pass2 = True

        if data_538[match]['team1'] == 'Bethlehem Steel FC':
            if data_538[match]['season'] == '2020':
                data_538[match]['team1'] = 'Philadelphia Union 2'
                Pass1 = True
        if data_538[match]['team2'] == 'Bethlehem Steel FC':
            if data_538[match]['season'] == '2020':
                data_538[match]['team2'] = 'Philadelphia Union 2'
                Pass2 = True

        if Pass1 == False:
            data_538[match]['team1'] = team_dict[data_538[match]['team1']]
        if Pass2 == False:
            data_538[match]['team2'] = team_dict[data_538[match]['team2']]

        data_538[match]['win1'] = ''
        data_538[match]['win2'] = ''
        data_538[match]['tie'] = ''

        data_538[match]['match_day'] = date_regex.search(data_538[match]['date']).group(3)
        data_538[match]['match_month'] = date_regex.search(data_538[match]['date']).group(2)
        data_538[match]['match_year'] = date_regex.search(data_538[match]['date']).group(1)

        del data_538[match]['xg1'], data_538[match]['xg2'], data_538[match]['score1'], \
            data_538[match]['score2'], data_538[match]['nsxg1'], data_538[match]['nsxg2'], \
            data_538[match]['adj_score1'], data_538[match]['adj_score2'], data_538[match]['date'], \
            data_538[match]['season'], data_538[match]['league']

        found = False
        for match2 in data_oddsportal:
            found = False
            days = [float(data_oddsportal[match2]['match_day'])]
            months = [float(data_oddsportal[match2]['match_month'])]
            if 30 in days or 31 in days or 28 in days or 29 in days:
                days.append(1)
                days.append(float(data_oddsportal[match2]['match_day']) - 1)
                if 12 in months:
                    months.append(1)
                else:
                    months.append(float(data_oddsportal[match2]['match_month']) + 1)
            elif 1 in days:
                days.append(31)
                days.append(30)
                days.append(28)
                days.append(29)
                days.append(2)
                if 1 in months:
                    months.append(12)
                else:
                    months.append(float(data_oddsportal[match2]['match_month']) - 1)
            else:
                days.append(float(data_oddsportal[match2]['match_day']) - 1)
                days.append(float(data_oddsportal[match2]['match_day']) + 1)
                days.append(float(data_oddsportal[match2]['match_day']) - 2)
                days.append(float(data_oddsportal[match2]['match_day']) + 2)
                days.append(float(data_oddsportal[match2]['match_day']) - 3)
                days.append(float(data_oddsportal[match2]['match_day']) + 3)
            if float(data_538[match]['match_day']) in days and \
                    float(data_538[match]['match_month']) in months and \
                    float(data_538[match]['match_year']) == float(data_oddsportal[match2]['match_year']) and \
                    ((data_538[match]['team1'] == data_oddsportal[match2]['team1'] and
                      data_538[match]['team2'] == data_oddsportal[match2]['team2']) or
                     (data_538[match]['team1'] == data_oddsportal[match2]['team2'] and
                      data_538[match]['team2'] == data_oddsportal[match2]['team1'])):
                data_538[match]['date'] = str(float(data_oddsportal[match2]['match_year']) +
                                              float(data_oddsportal[match2]['match_month']) / 12 +
                                              float(data_oddsportal[match2]['match_day']) / 30.44 / 12)
                data_538[match]['dayOfYear'] = str(float(data_oddsportal[match2]['match_month']) * 30.44 +
                                                   float(data_oddsportal[match2]['match_day']))
                data_538[match]['dayOfWeek'] = str(date(
                    int(data_oddsportal[match2]['match_year']),
                    int(data_oddsportal[match2]['match_month']),
                    int(data_oddsportal[match2]['match_day'])).weekday())
                data_538[match]['numberBookmarkers'] = data_oddsportal[match2]['numberBookmarkers']
                tzone = tf.timezone_at(lng=float(data_538[match]['stadium_lon']),
                                       lat=float(data_538[match]['stadium_lat']))
                gametime = float(data_oddsportal[match2]['matchTime']) + 1 + float(timezone_dict[tzone])
                if gametime < 0:
                    gametime += 24
                elif gametime > 24:
                    gametime -= 24
                data_538[match]['matchTime'] = str(gametime)
                data_538[match]['oddsW1'] = data_oddsportal[match2]['oddsW1']
                data_538[match]['oddsTie'] = data_oddsportal[match2]['oddsTie']
                data_538[match]['oddsW2'] = data_oddsportal[match2]['oddsW2']

                del data_oddsportal[match2]
                found = True
                break
        if found == False:
            del data_538[match]

    league_list = sorted(list(league_to_url_dictionary('oddsportal').keys()))

    for match in data_538:
        for league in league_list:
            if data_538[match]['league_id'] == league:
                data_538[match][str(league)] = '1'
            else:
                data_538[match][str(league)] = '0'
        del data_538[match]['league_id']

    jsonFile = json.dumps(data_538)
    f = open("merged_data.json", "w")
    f.write(jsonFile)
    f.close()
    json_to_excel()
