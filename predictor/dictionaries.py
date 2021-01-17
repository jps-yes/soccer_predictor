import openpyxl
import sys


# imports data from "stadiums.xlsx" returning the following dictionaries: stadium_capacity, stadium_lat, stadium_lon
# dictionary keys are team names
def stadium_dictionary():
    workbook = openpyxl.load_workbook(r'.\stadiums.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    y = 1
    stadium_capacity = {}
    stadium_lat = {}
    stadium_lon = {}
    while sheet.cell(row=y, column=1).value is not None:
        stadium_capacity[sheet.cell(row=y, column=1).value] = sheet.cell(row=y, column=2).value
        stadium_lat[sheet.cell(row=y, column=1).value] = sheet.cell(row=y, column=3).value
        stadium_lon[sheet.cell(row=y, column=1).value] = sheet.cell(row=y, column=4).value
        y += 1
    workbook.close()
    return stadium_capacity, stadium_lat, stadium_lon


# checks if scraped team exists in team_dict.xlsx and stadiums.xlsx files
def team_exists(team_names):
    team_dict = team_dictionary()
    teams = []
    new_team = False
    file = open('team_dict.txt', 'a')
    if isinstance(team_names, list):
        teams = teams + team_names
    elif isinstance(team_names, str):
        teams.append(team_names)
    for team in teams:
        if team not in team_dict:
            new_team = True
            file.write(team + '\n')
    file.close()
    if new_team:
        sys.exit("Teams in team_dict.txt need to be added to team_dict.xlsx and stadiums.xlsx")


# imports data from "team_dict.xlsx". Team name equivalence from 538 and oddsportal
def team_dictionary():
    workbook = openpyxl.load_workbook('team_dict.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    y = 1
    team_dict = {}
    while sheet.cell(row=y, column=1).value is not None:
        team_dict[sheet.cell(row=y, column=1).value] = sheet.cell(row=y, column=2).value
        y += 1
    workbook.close()
    return team_dict


def timezone_dictionary():
    workbook = openpyxl.load_workbook(r'.\timezones.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    y = 1
    time_dict = {}
    while sheet.cell(row=y, column=1).value is not None:
        time_dict[sheet.cell(row=y, column=1).value] = sheet.cell(row=y, column=2).value
        y += 1
    workbook.close()
    return time_dict


def league_to_url_dictionary(website):
    workbook = openpyxl.load_workbook(r'.\league_url.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    y = 1
    league_dict = {}
    if website == 'oddsportal':
        col = 2
    elif website == 'luckia':
        col = 3
    elif website == 'betano':
        col = 4
    while sheet.cell(row=y, column=1).value is not None:
        league_dict[str(sheet.cell(row=y, column=1).value)] = sheet.cell(row=y, column=col).value
        y += 1
    workbook.close()
    return league_dict
